# booking/views_export.py
# วางไว้ที่: booking/views_export.py
# เพิ่มใน urls.py:
#   from .views_export import export_excel
#   path('export/excel/', export_excel),

import io
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

from .models import (
    User, Building, Room, RoomFacility,
    Booking, BookingLog, DemandForecast,
    Notification, RoomUsageStat
)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_excel(request):
    """
    Export ข้อมูลทั้งหมดใน DB เป็น Excel
    GET /api/export/excel/?sheets=all   (default)
    GET /api/export/excel/?sheets=bookings,users
    """
    if request.user.role not in ['admin', 'staff']:
        from rest_framework.response import Response
        return Response({'error': 'จำกัดสิทธิ์เฉพาะผู้ดูแลระบบ'}, status=403)

    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        from rest_framework.response import Response
        return Response({'error': 'กรุณาติดตั้ง openpyxl: pip install openpyxl'}, status=500)

    # ─── Styles ───────────────────────────────────────────
    HEADER_FILL  = PatternFill('solid', start_color='1F4E79')
    HEADER_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ALT_FILL  = PatternFill('solid', start_color='D6E4F0')
    CELL_FONT = Font(name='Arial', size=10)
    CELL_ALIGN = Alignment(vertical='center')

    THIN = Side(style='thin', color='BBBBBB')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    TITLE_FONT = Font(bold=True, color='1F4E79', name='Arial', size=13)

    def style_header(ws, headers, row=2):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font  = HEADER_FONT
            cell.fill  = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = BORDER

    def style_row(ws, row_num, values, alt=False):
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font  = CELL_FONT
            cell.alignment = CELL_ALIGN
            cell.border = BORDER
            if alt:
                cell.fill = ALT_FILL

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    def add_title(ws, title, subtitle=''):
        ws.merge_cells('A1:Z1')
        ws['A1'] = title
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 28

    now_str = timezone.now().strftime('%d/%m/%Y %H:%M')
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # ลบ default sheet

    requested = request.query_params.get('sheets', 'all')

    def should_include(name):
        return requested == 'all' or name in requested.split(',')

    # ══════════════════════════════════════════════════════
    # Sheet 1: Users
    # ══════════════════════════════════════════════════════
    if should_include('users'):
        ws = wb.create_sheet('ผู้ใช้งาน')
        add_title(ws, f'ข้อมูลผู้ใช้งาน — Export {now_str}')
        headers = ['ID', 'Username', 'ชื่อ-นามสกุล', 'Email', 'Role',
                   'คณะ/หน่วยงาน', 'เบอร์โทร', 'สถานะ', 'วันที่สร้าง']
        style_header(ws, headers)
        users = User.objects.all().order_by('id')
        for i, u in enumerate(users):
            style_row(ws, i + 3, [
                u.id, u.username, u.get_full_name(),
                u.email, u.get_role_display() if hasattr(u, 'get_role_display') else u.role,
                u.faculty, u.phone,
                'Active' if u.is_active else 'Inactive',
                u.created_at.strftime('%d/%m/%Y %H:%M') if u.created_at else '',
            ], alt=(i % 2 == 1))
        auto_width(ws)
        ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════
    # Sheet 2: Buildings
    # ══════════════════════════════════════════════════════
    if should_include('buildings'):
        ws = wb.create_sheet('อาคาร')
        add_title(ws, f'ข้อมูลอาคาร — Export {now_str}')
        headers = ['ID', 'รหัสอาคาร', 'ชื่ออาคาร', 'คำอธิบาย', 'สถานะ', 'จำนวนห้อง']
        style_header(ws, headers)
        buildings = Building.objects.all().order_by('code')
        for i, b in enumerate(buildings):
            style_row(ws, i + 3, [
                b.id, b.code, b.name, b.description,
                'เปิดใช้งาน' if b.is_active else 'ปิดใช้งาน',
                b.rooms.count(),
            ], alt=(i % 2 == 1))
        auto_width(ws)
        ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════
    # Sheet 3: Rooms
    # ══════════════════════════════════════════════════════
    if should_include('rooms'):
        ws = wb.create_sheet('ห้อง')
        add_title(ws, f'ข้อมูลห้อง — Export {now_str}')
        headers = ['ID', 'ชื่อห้อง', 'อาคาร', 'รหัสอาคาร', 'ชั้น',
                   'ความจุ (คน)', 'ประเภทห้อง', 'สถานะ', 'สถานะการใช้งาน', 'วันที่สร้าง']
        style_header(ws, headers)
        rooms = Room.objects.select_related('building').order_by('building__code', 'name')
        for i, r in enumerate(rooms):
            style_row(ws, i + 3, [
                r.id, r.name, r.building.name, r.building.code,
                r.floor, r.capacity, r.room_type,
                r.get_status_display(),
                'เปิดใช้งาน' if r.is_active else 'ปิดใช้งาน',
                r.created_at.strftime('%d/%m/%Y') if r.created_at else '',
            ], alt=(i % 2 == 1))
        auto_width(ws)
        ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════
    # Sheet 4: Room Facilities
    # ══════════════════════════════════════════════════════
    if should_include('facilities'):
        ws = wb.create_sheet('อุปกรณ์ในห้อง')
        add_title(ws, f'อุปกรณ์ในห้อง — Export {now_str}')
        headers = ['ID', 'ห้อง', 'อาคาร', 'ชื่ออุปกรณ์', 'จำนวน']
        style_header(ws, headers)
        facilities = RoomFacility.objects.select_related('room__building').order_by('room__name')
        for i, f in enumerate(facilities):
            style_row(ws, i + 3, [
                f.id, f.room.name, f.room.building.name,
                f.name, f.quantity,
            ], alt=(i % 2 == 1))
        auto_width(ws)
        ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════
    # Sheet 5: Bookings
    # ══════════════════════════════════════════════════════
    if should_include('bookings'):
        ws = wb.create_sheet('การจอง')
        add_title(ws, f'ข้อมูลการจอง — Export {now_str}')
        headers = ['ID', 'ผู้จอง', 'Username', 'ห้อง', 'อาคาร',
                   'หัวข้อ', 'จำนวนคน', 'วันเริ่ม', 'เวลาเริ่ม', 'เวลาสิ้นสุด',
                   'สถานะ', 'Check-in', 'แจ้งเตือนแล้ว', 'อนุมัติโดย',
                   'หมายเหตุ', 'วันที่สร้าง']
        style_header(ws, headers)
        bookings = Booking.objects.select_related(
            'user', 'room__building', 'approved_by'
        ).order_by('-created_at')
        for i, b in enumerate(bookings):
            style_row(ws, i + 3, [
                b.id,
                b.user.get_full_name() or b.user.username,
                b.user.username,
                b.room.name, b.room.building.name,
                b.title, b.attendees,
                b.start_time.strftime('%d/%m/%Y'),
                b.start_time.strftime('%H:%M'),
                b.end_time.strftime('%H:%M'),
                b.get_status_display(),
                'ใช่' if b.checked_in else 'ไม่',
                'ใช่' if b.reminded else 'ไม่',
                b.approved_by.get_full_name() if b.approved_by else '',
                b.note,
                b.created_at.strftime('%d/%m/%Y %H:%M'),
            ], alt=(i % 2 == 1))
        auto_width(ws)
        ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════
    # Sheet 6: Booking Logs
    # ══════════════════════════════════════════════════════
    if should_include('logs'):
        ws = wb.create_sheet('ประวัติการจอง')
        add_title(ws, f'ประวัติการเปลี่ยนสถานะ — Export {now_str}')
        headers = ['ID', 'Booking ID', 'ผู้จอง', 'ห้อง', 'สถานะเดิม',
                   'สถานะใหม่', 'เปลี่ยนโดย', 'หมายเหตุ', 'วันที่เปลี่ยน']
        style_header(ws, headers)
        logs = BookingLog.objects.select_related(
            'booking__user', 'booking__room', 'changed_by'
        ).order_by('-changed_at')
        for i, l in enumerate(logs):
            style_row(ws, i + 3, [
                l.id, l.booking.id,
                l.booking.user.get_full_name() or l.booking.user.username,
                l.booking.room.name,
                l.old_status, l.new_status,
                l.changed_by.get_full_name() if l.changed_by else 'ระบบ',
                l.remark,
                l.changed_at.strftime('%d/%m/%Y %H:%M'),
            ], alt=(i % 2 == 1))
        auto_width(ws)
        ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════
    # Sheet 7: Demand Forecast
    # ══════════════════════════════════════════════════════
    if should_include('forecasts'):
        ws = wb.create_sheet('ผลพยากรณ์ AI')
        add_title(ws, f'ผลพยากรณ์ความต้องการ (AI) — Export {now_str}')
        headers = ['ID', 'ห้อง', 'อาคาร', 'วันที่พยากรณ์', 'ชั่วโมง',
                   'ค่าพยากรณ์ (0-1)', 'ระดับ', 'สถานะที่แสดง', 'ความมั่นใจ (%)']
        style_header(ws, headers)
        forecasts = DemandForecast.objects.select_related(
            'room__building'
        ).order_by('forecast_date', 'hour', 'room')
        for i, f in enumerate(forecasts):
            style_row(ws, i + 3, [
                f.id, f.room.name, f.room.building.name,
                f.forecast_date.strftime('%d/%m/%Y'),
                f'{f.hour:02d}:00',
                round(f.predicted_demand, 4),
                f.get_demand_level_display(),
                f.get_availability_display(),
                round(f.confidence, 1),
            ], alt=(i % 2 == 1))
        auto_width(ws)
        ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════
    # Sheet 8: Notifications
    # ══════════════════════════════════════════════════════
    if should_include('notifications'):
        ws = wb.create_sheet('การแจ้งเตือน')
        add_title(ws, f'ข้อมูลการแจ้งเตือน — Export {now_str}')
        headers = ['ID', 'ผู้รับ', 'Username', 'ประเภท', 'หัวข้อ',
                   'ข้อความ', 'อ่านแล้ว', 'Booking ID', 'วันที่']
        style_header(ws, headers)
        notifs = Notification.objects.select_related(
            'user', 'booking'
        ).order_by('-created_at')
        for i, n in enumerate(notifs):
            style_row(ws, i + 3, [
                n.id,
                n.user.get_full_name() or n.user.username,
                n.user.username,
                n.get_type_display(),
                n.title, n.message,
                'อ่านแล้ว' if n.is_read else 'ยังไม่อ่าน',
                n.booking.id if n.booking else '',
                n.created_at.strftime('%d/%m/%Y %H:%M'),
            ], alt=(i % 2 == 1))
        auto_width(ws)
        ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════
    # Sheet 9: Room Usage Stats
    # ══════════════════════════════════════════════════════
    if should_include('stats'):
        ws = wb.create_sheet('สถิติห้อง')
        add_title(ws, f'สถิติการใช้งานห้องรายวัน — Export {now_str}')
        headers = ['ID', 'ห้อง', 'อาคาร', 'วันที่', 'จองทั้งหมด',
                   'ใช้งานจริง', 'No-Show', 'ยกเลิก', 'อัตราการใช้งาน (%)']
        style_header(ws, headers)
        stats = RoomUsageStat.objects.select_related(
            'room__building'
        ).order_by('-date', 'room')
        for i, s in enumerate(stats):
            style_row(ws, i + 3, [
                s.id, s.room.name, s.room.building.name,
                s.date.strftime('%d/%m/%Y'),
                s.total_bookings, s.completed,
                s.no_show, s.cancelled,
                round(s.utilization_rate, 1),
            ], alt=(i % 2 == 1))
        auto_width(ws)
        ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════
    # Sheet 10: Summary (หน้าแรก)
    # ══════════════════════════════════════════════════════
    ws_sum = wb.create_sheet('สรุปภาพรวม', 0)
    ws_sum['A1'] = 'สรุปข้อมูลระบบจองห้องประชุม'
    ws_sum['A1'].font = Font(bold=True, size=16, color='1F4E79', name='Arial')
    ws_sum['A2'] = f'Export วันที่: {now_str}  |  Export โดย: {request.user.get_full_name() or request.user.username}'
    ws_sum['A2'].font = Font(size=10, color='666666', name='Arial', italic=True)

    ws_sum.merge_cells('A1:D1')
    ws_sum.merge_cells('A2:D2')
    ws_sum.row_dimensions[1].height = 30

    summary_data = [
        ('', '', '', ''),
        ('ตาราง', 'จำนวน Record', 'หมายเหตุ', ''),
        ('ผู้ใช้งาน (Users)',        User.objects.count(),            'ทุก Role', ''),
        ('อาคาร (Buildings)',        Building.objects.count(),         '', ''),
        ('ห้อง (Rooms)',             Room.objects.count(),             'ทุกสถานะ', ''),
        ('อุปกรณ์ (Facilities)',     RoomFacility.objects.count(),     '', ''),
        ('การจอง (Bookings)',        Booking.objects.count(),          'ทุกสถานะ', ''),
        ('  - กำลังจอง',            Booking.objects.filter(status='approved').count(),  '', ''),
        ('  - เสร็จสิ้น',           Booking.objects.filter(status='completed').count(), '', ''),
        ('  - No-Show',             Booking.objects.filter(status='no_show').count(),   '', ''),
        ('  - ยกเลิก',              Booking.objects.filter(status='cancelled').count(), '', ''),
        ('ประวัติการจอง (Logs)',     BookingLog.objects.count(),       '', ''),
        ('ผลพยากรณ์ AI (Forecasts)', DemandForecast.objects.count(),  '', ''),
        ('การแจ้งเตือน (Notifications)', Notification.objects.count(), '', ''),
        ('สถิติห้อง (Stats)',        RoomUsageStat.objects.count(),    '', ''),
    ]

    for r, row_data in enumerate(summary_data, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=r, column=c, value=val)
            cell.font = Font(name='Arial', size=10,
                             bold=(row_data[0] == 'ตาราง'),
                             color='FFFFFF' if row_data[0] == 'ตาราง' else '000000')
            if row_data[0] == 'ตาราง':
                cell.fill = PatternFill('solid', start_color='1F4E79')
            elif r % 2 == 0:
                cell.fill = PatternFill('solid', start_color='D6E4F0')
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center',
                                       horizontal='right' if c == 2 else 'left')

    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 18
    ws_sum.column_dimensions['C'].width = 20

    # ─── Response ─────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'room_booking_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
